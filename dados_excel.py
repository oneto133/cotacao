import openpyxl
import pandas as pd
from Funções import Tempo, Arquivo
import shutil
import win32com.client as win32



class main(Tempo, Arquivo):
    def __init__(self):
        super().__init__()
        self.origem = r"C:\Users\rodri\OneDrive\Banco Gênio\Banco Gênio.xlsx"
        self.destino = r"C:\Users\rodri\OneDrive\Banco Gênio\Banco_gênio_temp.xlsx"
        self.base = r"C:\Users\rodri\OneDrive\Banco Gênio\Base.xlsx"
        self.arquivos()

    def arquivos(self):
        try:
            df = pd.read_csv(r'csv/indice.csv')
            shutil.copy2(self.origem, self.destino)
            arq = openpyxl.load_workbook(self.destino, data_only=True)
            base = openpyxl.load_workbook(self.base)
            self.programa(df, arq, base)
        except PermissionError as e:
            print("Parece que o arquivo está aberto, vamos tentar fechar para você")
            try:
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                pasta = excel.Workbooks.Open(self.origem)
                pasta.Close(SaveChanges=False)
                pasta = excel.Workbooks.Open(self.base)
                pasta.Close(SaveChanges=True)
                excel.Quit()
            except Exception as e:
                print("Erro ao tenar fechar a pasta: ", e)




    def programa(self, df, arq, base):

        linha_destino = int(df.columns[0])
        coluna_destino = 1
        Financeiro = arq["Financeiro"]
        Banco_de_dados = base["Base"]
        Intervalo = ["A5:G5", "J6:Q6"]
        Dados = []

        for intervalo in Intervalo:
            for row in Financeiro[intervalo]:
                dado = [cell.value for cell in row]
                Dados.extend(dado)

        for coluna, valor in enumerate(Dados):
            Banco_de_dados.cell(row=linha_destino, column=coluna_destino + coluna, value=valor)

        Banco_de_dados.cell(row=linha_destino, column=coluna+2, value=self.data_atual())
        Banco_de_dados.cell(row=linha_destino, column=coluna+3, value=self.hora_atual())

        with open("csv/indice.csv", "w") as file:
            file.write(f'"{linha_destino+1}",{self.data_atual()},{self.hora_atual()}')

        base.save(r"C:\Users\rodri\OneDrive\Banco Gênio\Base.xlsx")
        self.excluir(self.destino)


if __name__ == "__main__":
    df = pd.read_csv(r"csv/indice.csv", encoding="latin1")
    data = df.columns[1]
    Hor = Tempo()
    dat = Hor.data_atual()
    if data != dat:
        main()
    else:
        print("Não executado!")